import argparse
import copy
import os
import typing as tp

import openpyxl
from datetime import date
from rich.console import Console

SOURCE_PATH = './source'
OUTPUT_PATH = './out'


def _get_files() -> list:
    if os.path.exists(SOURCE_PATH):
        all_files = os.listdir(SOURCE_PATH)
        return list(filter(lambda f: f.endswith('.xlsx'), all_files))
    return []


def _save_sheet_as_new_workbook(workbook: openpyxl.Workbook, workbook_name: str, sheet_name: str,
                                is_separately: bool = False,
                                with_filename_column: bool = False, filename_sheets: tp.Optional[tp.List[str]] = None,
                                filename_column_name: tp.Optional[str] = None):
    new_workbook = copy.deepcopy(workbook)
    sheets = workbook.sheetnames

    for sheet in sheets:
        if sheet == sheet_name:
            continue
        del new_workbook[sheet]

    if with_filename_column and len(filename_sheets):
        if sheet_name.lower() in filename_sheets:
            ws = new_workbook.active
            ws.insert_cols(0, 1)

            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4)):
                if row[1].value is None:
                    continue

                if i == 0:
                    row[0].value = filename_column_name if filename_column_name else 'Filename'
                else:
                    row[0].value = workbook_name

    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)

    today = date.today()
    file_path = '{}/{}'.format(OUTPUT_PATH, workbook_name)
    if is_separately:
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        path_to_save = '{}/{}_{}.xlsx'.format(file_path, sheet_name, today.strftime("%d_%m_%Y"))
    else:
        path_to_save = '{}_{}_{}.xlsx'.format(file_path, sheet_name, today.strftime("%d_%m_%Y"))
    new_workbook.save(path_to_save)


def _separate_file(file: str, is_separately: bool = False,
                   with_filename_column: bool = False, filename_sheets: tp.Optional[tp.List[str]] = None,
                   filename_column_name: tp.Optional[str] = None):
    file_path = '{}/{}'.format(SOURCE_PATH, file)
    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames

    workbook_name = file.split('.')[0]
    for sheet in sheets:
        _save_sheet_as_new_workbook(workbook, workbook_name, sheet, is_separately,
                                    with_filename_column, filename_sheets, filename_column_name)


def run():
    parser = argparse.ArgumentParser(description='Separate workbook sheets as self workbooks.')
    parser.add_argument('--by-folders', dest='is_separately', action='store_const',
                        const=True, default=False,
                        help='Out workbooks separately by folders (default: place all results to out directory)')
    parser.add_argument('--with-filename-column', dest='with_filename_column', action='store_const',
                        const=True, default=False,
                        help='Add new column with filename to child sheets')
    parser.add_argument('--filename-sheets', nargs='+',
                        help='Add new column with filename to child sheets')
    parser.add_argument('--filename-column-name', type=str,
                        help='Filename column name')
    args = parser.parse_args()

    console = Console()

    source_files = _get_files()
    count_files = len(source_files)
    console.log(f"[bold yellow]Files to process: {count_files}")

    with console.status("[bold green]Fetching data...") as status:
        counter = 0
        for file in source_files:
            counter += 1
            console.log(f"[green]Processing file[/green] {counter}: {file}")
            filename_sheets = [x.lower() for x in args.filename_sheets]
            _separate_file(file, args.is_separately,
                           args.with_filename_column, filename_sheets, args.filename_column_name)
        console.log(f'[bold][red]Done!')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    run()
